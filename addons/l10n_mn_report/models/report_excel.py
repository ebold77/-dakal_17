# -*- coding: utf-8 -*-
import json
import pytz
from datetime import datetime
import base64
import xlrd
import io
import logging

from odoo import models, fields, api
# from odoo.tools.translate import _, translate
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

if 'coordinate_from_string' not in globals():
    try:
        from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
    except ImportError:
        logger.warning('Cannot import openpyxl.utils!')


class ReportExcelOutput(models.TransientModel):
    _name = 'oderp.report.excel.output'
    _description = "Excel Report Output"

    filename = fields.Char('File name', readonly=True)
    filedata = fields.Binary('File data', readonly=True)

    @api.model
    def export_report(self):
        self.ensure_one()

        # get time
        if self.env.user.partner_id.tz:
            tz = pytz.timezone(self.env.user.partner_id.tz)
        else:
            tz = pytz.utc
        now_utc = datetime.now(pytz.timezone('UTC'))
        now_user_zone = now_utc.astimezone(tz)

        filename_prefix = self.env.context.get('filename_prefix', 'report_excel_output')
        filename = "%s_%s.xlsx" % (filename_prefix, now_user_zone.strftime('%Y%m%d_%H%M%S'))
        self.filename = filename

        form_title = self.env.context.get('form_title', _('Report Result'))

        mod_obj = self.env['ir.model.data']
        form_res = mod_obj.get_object_reference('l10n_mn_report', 'report_excel_output_view_form')
        form_id = form_res and form_res[1] or False
        if 'oderp_report_html' not in self.env.context:
            return {
                'name': form_title,
                'view_type': 'form',
                'view_mode': 'form',
                'res_model': 'oderp.report.excel.output',
                'res_id': self.id,
                'views': [(form_id, 'form')],
                'context': self.env.context,
                'type': 'ir.actions.act_window',
                'target': 'new',
            }
        else:
            return self

    # def _(self, source):
    #     '''
    #     Translation method for selection field
    #     '''
    #     # currently only in selection fields
    #     result = translate(self.env.cr, False, 'selection', self.env.context['lang'], source) or source
    #     return result

    def get_excel_file_value(self):
        filedata = self.filedata
        book = xlrd.open_workbook(file_contents=base64.decodestring(filedata))
        sheet = book.sheet_by_name(book.sheet_names()[0])
        data = [[sheet.cell_value(r, c) for c in range(sheet.ncols)] for r in range(sheet.nrows)]
        return json.dumps(data)

    def coordinate_to_string(self, coordinate):
        xy = coordinate_from_string(coordinate.partition(":")[0])
        col = column_index_from_string(xy[0])
        row = xy[1]
        xy = coordinate_from_string(coordinate.split(":", 1)[1])
        colspan = column_index_from_string(xy[0])
        rowspan = xy[1]
        return [row, col, (rowspan + 1) - row, (colspan + 1) - col]

    def get_excel_file_merge(self):
        wb = load_workbook(filename=io.BytesIO(base64.decodestring(self.filedata)))
        sheets = wb.sheetnames
        sheet_ranges = wb[sheets[0]]
        cell_ranges = []
        cell_range_cor = sheet_ranges.merged_cells.ranges
        for obj in cell_range_cor:
            coor_str = str(obj)
            range_arr = self.coordinate_to_string(coor_str)
            cell_ranges.append(range_arr)
        return json.dumps(cell_ranges)

    def get_emp_name_from_sign(self, sign, company=False):
        self.ensure_one()
        sign_user_id = False
        company = self.env.user.company_id if not company else company

        if sign == 'first' and company.first_sign_user_id:
            sign_user_id = company.first_sign_user_id
        elif sign == 'second' and company.second_sign_ser_id:
            sign_user_id = company.second_sign_ser_id
        elif sign == 'ceo' and company.executive_sign_user_id:
            sign_user_id = company.executive_sign_user_id
        elif sign == 'general_account' and company.general_accountant_sign_user_id:
            sign_user_id = company.general_accountant_sign_user_id
        elif sign == 'account' and company.accountant_sign_user_id:
            sign_user_id = company.accountant_sign_user_id

        if sign_user_id:
            sign_employee = self.env['hr.employee'].search([('company_id', '=', company.sudo().id), ('resource_id.user_id', '=', sign_user_id.id)], limit=1)
            if sign_employee:
                return sign_employee.sudo().name or '____________________'
            else:
                return sign_user_id.sudo().name or '____________________'
        else:
            return '____________________'